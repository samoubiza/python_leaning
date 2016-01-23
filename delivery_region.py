#-------------------------------------------------------------------------------
# Name:        модуль1
# Purpose:
#
# Author:      onovikov
#
# Created:     20.01.2016
# Copyright:   (c) onovikov 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------

def main():
    pass

if __name__ == '__main__':
    main()
import openpyxl as px
import numpy as np
from fuzzywuzzy import fuzz
from fuzzywuzzy import process


my_xl = px.load_workbook('C:\Temp\BI\TMS\python_test.xlsx')
our_page = my_xl.get_sheet_by_name(name = 'our')

dict_our={}

for row in our_page.iter_rows('A2:C50982'):
    l =[]
    for cell in row:
        l.append(cell.value)
    ##dict_our[l[0]]=",".join(l[1:])
    dict_our[l[0]]=l[1:]

del_page = my_xl.get_sheet_by_name(name = 'delivery')

dict_del={}

for row in del_page.iter_rows('A2:C5230'):
    l =[]
    for cell in row:
        l.append(cell.value)
    ##dict_del[l[0]]=",".join(l[1:])
    dict_del[l[0]]=l[1:]


multiple_max = {}
ourID_maxDELID = {}
for k in dict_our.keys():
    max_fuzzy={}


    dict_del_reg = dict((k_, v_) for k_, v_ in dict_del.items() if fuzz.token_set_ratio(str(v_[1]) ,str(dict_our[k][1])) >=95)

    if len(dict_del_reg)>0:
        for id in dict_del_reg.keys():
            fuzzy= fuzz.token_set_ratio(",".join(dict_our[k]),",".join(dict_del_reg[id]))
            if id in max_fuzzy.keys():
                max_fuzzy[id].append(fuzzy)
            else:
                max_fuzzy[id]=[]
                max_fuzzy[id].append(fuzzy)
        max_rate = [key for key,val in max_fuzzy.items() if val == max(max_fuzzy.values())]
        ourID_maxDELID[k]=max_rate[0]
        if len(max_rate)>1:
            multiple_max[k]=max_rate
    else:
        ourID_maxDELID[k]=0

##
out=open('C:/Temp/BI/TMS/result.txt','w')

for key,value in ourID_maxDELID.items():
    st = str(key)+";"+str(value)
    out.write(st)
    out.write('\n')

out.close()