#-------------------------------------------------------------------------------
# Name:        модуль1
# Purpose:
#
# Author:      onovikov
#
# Created:     22.01.2016
# Copyright:   (c) onovikov 2016
# Licence:     <your licence>
#-------------------------------------------------------------------------------

def main():
    pass

if __name__ == '__main__':
    main()

import openpyxl as px
import numpy as np
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

my_xl = px.load_workbook('C:\Temp\BI\TMS\python_test2.xlsx')
our_page = my_xl.get_sheet_by_name(name = 'our')

dict_our={}

for row in our_page.iter_rows('A2:D28'):
    l =[]
    for cell in row:
        l.append(cell.value)
    ##dict_our[l[0]]=",".join(l[1:])
    dict_our[l[0]]=l[1:]

del_page = my_xl.get_sheet_by_name(name = 'delivery')

dict_del={}

for row in del_page.iter_rows('A2:C38'):
    l =[]
    for cell in row:
        l.append(cell.value)
    ##dict_del[l[0]]=",".join(l[1:])
    dict_del[l[0]]=l[1:]

for k in dict_del.keys():
    current_reg = 0
    id_reg_reg = 0
    max_fuzzy={}
    current_reg = dict_del[k][1]

    dict_our_reg = dict((k_, v_) for k_, v_ in dict_our.items() if fuzz.token_set_ratio(str(v_[1]) ,str(current_reg)) >=95)
    dict_del_reg = dict((i, j) for i, j in dict_del.items() if fuzz.token_set_ratio(str(j[1]) ,str(current_reg)) >=95)
    current_reg_id = dict_our_reg[list(dict_our_reg)[0]][2]
    for id in dict_del_reg.keys():
        if fuzz.token_set_ratio(dict_del_reg[id][0],dict_del_reg[id][1]) ==100:
            id_reg_reg = id

    if k == id_reg_reg:
        dict_del[k].append(0)
        dict_del[k].append(current_reg_id)
    else:
        if len(dict_our_reg)>0:
            for key in dict_our_reg.keys():
                fuzzy= fuzz.token_set_ratio(str(dict_del[k][0]),str(dict_our_reg[key][0]))
                if key in max_fuzzy.keys() and fuzzy > 85:
                    max_fuzzy[key].append(fuzzy)
                elif fuzzy >85:
                    max_fuzzy[key]=[]
                    max_fuzzy[key].append(fuzzy)
            max_rate = [ke for ke,va in max_fuzzy.items() if va == max(max_fuzzy.values())]
            if len(max_rate)>0:
                dict_del[k].append(max_rate[0])
                dict_del[k].append(current_reg_id)
            else:
                dict_del[k].append(-1)
                dict_del[k].append(current_reg_id)

out=open('C:/Temp/BI/TMS/result2.txt','w')

for key,value in dict_del.items():
    st = str(key)+";"+str(value[0])+";"+str(value[1])+";"+str(value[2])+";"+str(value[3])
    out.write(st)
    out.write('\n')

out.close()